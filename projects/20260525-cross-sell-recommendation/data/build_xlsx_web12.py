#!/usr/bin/env python3
"""nbo_web12_named.csv → 공유용 Excel (웹 인기 12스킬 추천세트)."""
import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT='맑은 고딕'
HDR=PatternFill('solid', fgColor='C55A11'); ALT=PatternFill('solid', fgColor='FBE9DD')  # 웹=주황 계열로 앱셋과 구분
GREEN=Font(name=FONT,size=10,color='1F7A1F'); GRAY=Font(name=FONT,size=10,color='808080')
hf=Font(name=FONT,size=10,bold=True,color='FFFFFF'); bf=Font(name=FONT,size=10)
thin=Side(style='thin',color='D9D9D9'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal='center',vertical='center'); lft=Alignment(horizontal='left',vertical='center')

d=pd.read_csv('nbo_web12_named.csv')
d['frac']=d['score_pct']/100
WEBORDER=[59089,43117,49816,60706,55723,57670,21964,27376,47704,60013,2141,60244]
d['ord']=d.anchor.map({x:i for i,x in enumerate(WEBORDER)})
d=d.sort_values(['ord','rank'])

def header(ws,n):
    for c in range(1,n+1):
        cell=ws.cell(1,c); cell.font=hf; cell.fill=HDR; cell.alignment=ctr; cell.border=BORD
    ws.freeze_panes='A2'; ws.row_dimensions[1].height=22

wb=Workbook()

# 시트1: 추천세트(웹) wide
ws=wb.active; ws.title='추천세트(웹)'
ws.append(['웹 인기스킬(첫 구매)','주제','앱표본','데이터유형']+[f'추천{i}' for i in range(1,11)])
for i,X in enumerate(WEBORDER):
    g=d[d.anchor==X].sort_values('rank');
    if not len(g): continue
    recs=g['rec_name'].tolist()[:10]+['']*(10-len(g))
    typ='직접(앱)' if (g['source']=='직접').any() else '관심사(토픽)'
    row=[g['anchor_name'].iat[0],g['anchor_topic'].iat[0],int(g['anchor_n_app'].iat[0]),typ]+recs
    ws.append(row); r=i+2
    for c in range(1,15):
        cell=ws.cell(r,c); cell.font=bf; cell.border=BORD; cell.alignment=ctr if c in(2,3,4) else lft
        if i%2: cell.fill=ALT
    ws.cell(r,3).number_format='#,##0'; ws.cell(r,4).font=GREEN if typ=='직접(앱)' else GRAY
for c,w in zip('ABCD',[30,9,8,12]): ws.column_dimensions[c].width=w
for c in 'EFGHIJKLMN': ws.column_dimensions[c].width=26
header(ws,14)

# 시트2: 상세
ws2=wb.create_sheet('상세')
ws2.append(['웹스킬코드','웹 인기스킬','주제','앱표본','순위','추천코드','추천 스킬','출처','함께/선호율','lift'])
for i,(_,r) in enumerate(d.iterrows()):
    lift='' if pd.isna(r.lift) else float(r.lift)
    ws2.append([int(r.anchor),r.anchor_name,r.anchor_topic,int(r.anchor_n_app),int(r['rank']),int(r.rec_menu),r.rec_name,r.source,r.frac,lift])
    rr=i+2
    for c in range(1,11):
        cell=ws2.cell(rr,c); cell.font=bf; cell.border=BORD; cell.alignment=ctr if c in(1,4,5,6,8,9,10) else lft
        if i%2: cell.fill=ALT
    ws2.cell(rr,4).number_format='#,##0'; ws2.cell(rr,9).number_format='0.0%'
    if lift!='': ws2.cell(rr,10).number_format='0.0"x"'
    ws2.cell(rr,8).font=GREEN if r.source=='직접' else GRAY
for c,w in zip('ABCDEFGHIJ',[10,30,9,8,6,9,30,9,11,7]): ws2.column_dimensions[c].width=w
header(ws2,10)

# 시트0: 읽는 법
ws3=wb.create_sheet('읽는 법',0)
ws3.column_dimensions['A'].width=30; ws3.column_dimensions['B'].width=80
def put(r,c,v,bold=False,size=10,color='000000'):
    cell=ws3.cell(r,c,v); cell.font=Font(name=FONT,size=size,bold=bold,color=color)
    cell.alignment=Alignment(horizontal='left',vertical='center',wrap_text=(c==2)); return cell
put(1,1,'헬로우봇 — 웹 인기 12스킬 추천세트',bold=True,size=14,color='C55A11')
put(2,1,'웹 스토어에서 많이 팔리는 12개 스킬을 산 사용자에게 보여줄 추천 (앱 행동으로 관심사 추정)',color='808080')
rows=[
 ('■ 핵심 아이디어',''),
 ('왜 앱 데이터로?','웹 재구매는 마케팅 캠페인(끼워팔기)에 좌우돼 "관심"을 왜곡. 그래서 같은 스킬을 산 *앱* 사용자가 함께 무엇에 관심 갖는지로 추정.'),
 ('','' ),
 ('■ 2층 방식',''),
 ('1층 직접(앱)','그 스킬을 산 앱 사용자가 함께 많이 산 스킬 (양방향 함께구매, 안정화 정렬). 데이터 충분한 9개.'),
 ('2층 관심사(토픽)','앱 신호가 거의 없는 웹획득형 스킬은 주제 그룹의 선호 스킬로 채움.'),
 ('웹획득형 3개','결혼 OX 진단 · 내 운이 트이는 동네 · 연애 상대 스포 → 앱 구매 4~8건뿐(주제 미태깅) = 순수 웹유입. 이름으로 주제 보정 후 토픽 추천.'),
 ('위생','판매중·유료·콘텐츠만 추천, 자기 자신 제외. 보유 스킬 제외는 노출 시점.'),
 ('','' ),
 ('■ 읽는 법',''),
 ("'추천세트(웹)'",'한 줄=웹 인기스킬 1개. 추천1~10. 데이터유형=직접(앱)/관심사(토픽).'),
 ("'상세'",'순위·출처·함께율·lift. 함께/선호율=그 스킬 구매자 중 추천을 함께 산 비율(직접) 또는 주제 선호율(관심사). 초록=직접 / 회색=토픽.'),
 ('','' ),
 ('■ 주의',''),
 ('','추천은 "앱에서의 관심 연관"이지 효과 보장이 아님. 웹 사용자에 대한 실제 효과는 노출 vs 비노출 실험으로 별도 확정.'),
]
r=4
for a,b in rows:
    bold=a.startswith('■'); put(r,1,a,bold=bold,color='C55A11' if bold else '000000'); put(r,2,b); r+=1
ws3.freeze_panes='A2'

wb.save('nbo_web12_recommendation_set.xlsx')
print('saved nbo_web12_recommendation_set.xlsx ·', d.anchor.nunique(),'웹앵커 ·',len(d),'행')
