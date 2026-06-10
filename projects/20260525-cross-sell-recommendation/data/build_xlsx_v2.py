#!/usr/bin/env python3
"""nbo_v2_combined.csv → 공유용 Excel (추천세트 요약 / 상세 / 읽는 법)."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT='맑은 고딕'
HDR=PatternFill('solid', fgColor='2F5496'); ALT=PatternFill('solid', fgColor='F2F2F2')
GREEN=Font(name=FONT,size=10,color='1F7A1F'); GRAY=Font(name=FONT,size=10,color='808080')
hf=Font(name=FONT,size=10,bold=True,color='FFFFFF'); bf=Font(name=FONT,size=10)
thin=Side(style='thin',color='D9D9D9'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal='center',vertical='center'); lft=Alignment(horizontal='left',vertical='center')

d=pd.read_csv('nbo_v2_combined.csv')
d[['주제','세부관심']]=d['anchor_topic'].str.split('|',n=1,expand=True)
d['세부관심']=d['세부관심'].replace('*','-').fillna('-')
d['frac']=d['score_pct']/100
d=d.sort_values(['anchor_n','anchor','rank'],ascending=[False,True,True])

def style_header(ws,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(1,c); cell.font=hf; cell.fill=HDR; cell.alignment=ctr; cell.border=BORD
    ws.freeze_panes='A2'; ws.row_dimensions[1].height=22

wb=Workbook()

# ---------- 시트1: 추천세트(요약, wide) ----------
ws=wb.active; ws.title='추천세트'
cols=['첫 스킬(구매)','구매자수','주제','세부관심','데이터유형','추천1','추천2','추천3','추천4','추천5']
ws.append(cols)
wide=[]
for X,g in d.groupby('anchor',sort=False):
    g=g.sort_values('rank'); recs=g['rec_name'].tolist()[:5]+['']*(5-len(g))
    typ='직접 보유' if (g['source']=='직접').any() else '관심사 기반'
    wide.append([g['anchor_name'].iat[0],int(g['anchor_n'].iat[0]),g['주제'].iat[0],g['세부관심'].iat[0],typ]+recs)
wide=sorted(wide,key=lambda r:-r[1])
for i,row in enumerate(wide):
    ws.append(row); r=i+2
    for c in range(1,11):
        cell=ws.cell(r,c); cell.font=bf; cell.border=BORD
        cell.alignment=ctr if c in(2,3,4,5) else lft
        if i%2: cell.fill=ALT
    ws.cell(r,2).number_format='#,##0'
    ws.cell(r,5).font=GREEN if row[4]=='직접 보유' else GRAY
for c,w in zip('ABCDEFGHIJ',[32,9,10,10,11,28,28,28,28,28]): ws.column_dimensions[c].width=w
style_header(ws,10)

# ---------- 시트2: 상세(long) ----------
ws2=wb.create_sheet('상세')
cols2=['첫스킬코드','첫 스킬','구매자수','주제','세부관심','순위','추천스킬코드','추천 스킬','출처','전환율']
ws2.append(cols2)
for i,(_,r) in enumerate(d.iterrows()):
    ws2.append([int(r.anchor),r.anchor_name,int(r.anchor_n),r['주제'],r['세부관심'],int(r['rank']),
                int(r.rec_menu),r.rec_name,r.source,r.frac])
    rr=i+2
    for c in range(1,11):
        cell=ws2.cell(rr,c); cell.font=bf; cell.border=BORD
        cell.alignment=ctr if c in(1,3,6,7,9,10) else lft
        if i%2: cell.fill=ALT
    ws2.cell(rr,3).number_format='#,##0'; ws2.cell(rr,10).number_format='0.0%'
    ws2.cell(rr,9).font=GREEN if r.source=='직접' else GRAY
for c,w in zip('ABCDEFGHIJ',[11,32,9,10,10,6,11,32,8,9]): ws2.column_dimensions[c].width=w
style_header(ws2,10)

# ---------- 시트3: 읽는 법 ----------
ws3=wb.create_sheet('읽는 법',0)
def put(r,c,v,bold=False,size=10,color='000000',fill=None,wrap=False):
    cell=ws3.cell(r,c,v); cell.font=Font(name=FONT,size=size,bold=bold,color=color)
    cell.alignment=Alignment(horizontal='left',vertical='center',wrap_text=wrap)
    if fill: cell.fill=PatternFill('solid',fgColor=fill)
    return cell
ws3.column_dimensions['A'].width=30; ws3.column_dimensions['B'].width=78
put(1,1,'헬로우봇 — 결제 직후 재구매 추천 세트 (v2)',bold=True,size=14,color='2F5496')
put(2,1,'첫 스킬을 산 사용자에게 그 자리(결제 직후 슬롯)에서 보여줄 2번째 추천 스킬',color='808080')
rows=[
 ('■ 무엇인가',''),
 ('한 줄','첫 스킬 A를 산 사람 → 추천할 B 스킬 top5. "A 결제 → A 행의 추천1~5를 결제직후 슬롯에 노출".'),
 ('데이터','앱퍼스트 진짜 첫구매자 98,093명의 실제 1→2번째 구매(당일 인세션 포함).'),
 ('','' ),
 ('■ 어떻게 만들었나 (2층)',''),
 ('1층 직접','A를 산 사람들이 실제로 2번째로 많이 산 B (표본 충분한 스킬). 안정화(Wilson) 정렬.'),
 ('2층 관심사 백오프','직접 데이터가 부족하면 → A의 관심사 그룹(주제+세부관심)에서 사람들이 선호(전환율 높은)하는 스킬로 채움.'),
 ('위생 규칙','판매중·유료·콘텐츠 스킬만 추천. 방금 산 그 스킬·충전/질문권 등 제외. (이미 보유한 다른 스킬 제외는 노출 시점에 적용)'),
 ('','' ),
 ('■ 어떻게 읽나',''),
 ("'추천세트' 시트",'한 줄 = 첫 스킬 1개. 추천1~5가 보여줄 스킬. 데이터유형=직접 보유/관심사 기반.'),
 ("'상세' 시트",'순위·출처(직접/관심사)·전환율까지. 구현용. 전환율=그 스킬을 2번째로 산 비율.'),
 ('색 표시','초록=직접 데이터 / 회색=관심사 백오프.'),
 ('','' ),
 ('■ 커버리지 (자동 계산)',''),
]
r=4
for a,b in rows:
    bold=a.startswith('■'); put(r,1,a,bold=bold,color='2F5496' if bold else '000000'); put(r,2,b,wrap=True); r+=1
# 라이브 통계 (상세 시트 참조 수식)
stats=[('첫 스킬(앵커) 수',"=COUNTIF('상세'!F:F,1)",'0'),
       ('총 추천 행',"=COUNTA('상세'!A:A)-1",'0'),
       ('직접 데이터 보유 앵커',"=COUNTIFS('상세'!F:F,1,'상세'!I:I,\"직접\")",'0'),
       ('직접 추천 받는 첫구매자 비중',"=SUMIFS('상세'!C:C,'상세'!F:F,1,'상세'!I:I,\"직접\")/SUMIFS('상세'!C:C,'상세'!F:F,1)",'0.0%')]
for a,f,fmt in stats:
    put(r,1,a); cell=put(r,2,f); cell.number_format=fmt; cell.font=Font(name=FONT,size=10,bold=True); r+=1
r+=1
put(r,1,'■ 주의',bold=True,color='2F5496'); put(r+1,2,'추천 순서는 "현재 데이터의 연관"이지 효과 보장이 아님. 실제 매출 효과는 노출군 vs 비노출군 실험으로 별도 확정 필요.',wrap=True)
ws3.freeze_panes='A2'

wb.save('nbo_v2_recommendation_set.xlsx')
print('saved nbo_v2_recommendation_set.xlsx ·', d.anchor.nunique(),'anchors ·',len(d),'rows')
